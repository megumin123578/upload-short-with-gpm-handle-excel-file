import os
import itertools
import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
from hyperparameter import APP_TITLE, GROUPS_DIR, CHANNEL_HEADER_HINTS, OUTPUT_DIR, HAS_TKCAL
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tkcalendar import DateEntry

def list_group_csvs(groups_dir: str):
    if not os.path.isdir(groups_dir):
        return []
    files = []
    for name in os.listdir(groups_dir):
        p = os.path.join(groups_dir, name)
        nl = name.lower()
        if os.path.isfile(p) and nl.endswith(".csv"):
            if "__assignments_" in nl or nl.startswith("assignments_"):
                continue
            files.append(name)
    return sorted(files)

def read_channels_from_csv(csv_path: str):
    channels = []
    if not os.path.isfile(csv_path):
        return channels
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return channels
    header = [c.strip() for c in rows[0]] if rows and rows[0] else []
    col_idx = None
    if header:
        lower_header = [h.lower() for h in header]
        for hint in CHANNEL_HEADER_HINTS:
            if hint in lower_header:
                col_idx = lower_header.index(hint)
                break
    start_row = 1 if col_idx is not None else 0
    for row in rows[start_row:]:
        if not row:
            continue
        if col_idx is not None and col_idx < len(row):
            v = (row[col_idx] or "").strip()
            if v:
                channels.append(v)
        else:
            first = next((c.strip() for c in row if c and c.strip()), "")
            if first:
                channels.append(first)
    return channels

def normalize_lines(s: str):
    return [ln.strip() for ln in s.splitlines() if ln.strip()]

def assign_pairs(channels, titles, descs, mode="titles"):
    """
    mode = "titles": số dòng = len(titles); kênh & mô tả chạy vòng.
    mode = "channels": số dòng = len(channels); tiêu đề & mô tả chạy vòng.
    """
    if not channels:
        raise ValueError("No channels found in selected CSV.")
    if not titles:
        raise ValueError("No titles provided.")

    if len(descs) <= 1:
        desc_cycle = itertools.cycle([descs[0] if descs else ""])
    else:
        desc_cycle = itertools.cycle(descs)

    if mode == "titles":
        ch_cycle = itertools.cycle(channels)
        out = []
        for i, title in enumerate(titles):
            ch = next(ch_cycle)
            d = next(desc_cycle)
            out.append((ch, title, d))
        return out
    else:  # mode == "channels"
        title_cycle = itertools.cycle(titles)
        out = []
        for ch in channels:
            t = next(title_cycle)
            d = next(desc_cycle)
            out.append((ch, t, d))
        return out

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.state("zoomed")
        self.minsize(880, 580)

        self.group_file_var = tk.StringVar(value="")
        self.mode_var = tk.StringVar(value="titles")  # "titles" (mặc định) hoặc "channels"
        self.status_var = tk.StringVar(value="Ready.")
        self._channels_cache = []
        self._last_assignments = None

        self._build_header()
        self._build_inputs()
        self._build_preview()
        self._build_footer()
        self._refresh_group_files()

    def _build_header(self):
        frm = ttk.Frame(self, padding=(10, 10, 10, 0))
        frm.pack(fill=tk.X)

        ttk.Label(frm, text="Group CSV (in ./group):").grid(row=0, column=0, sticky="w")
        self.group_combo = ttk.Combobox(frm, textvariable=self.group_file_var, state="readonly", width=48)
        self.group_combo.grid(row=0, column=1, sticky="w", padx=6)
        self.group_combo.bind("<<ComboboxSelected>>", lambda e: self._load_channels())

        ttk.Button(frm, text="Refresh", command=self._refresh_group_files).grid(row=0, column=2, padx=4)
        ttk.Button(frm, text="Open group folder", command=self._open_group_dir).grid(row=0, column=3, padx=4)

        self.channel_count_lbl = ttk.Label(frm, text="0 channels")
        self.channel_count_lbl.grid(row=0, column=4, sticky="w", padx=(12, 0))
        ttk.Button(frm, text="📅 Date", width=8, command=self._pick_date_for_selection).grid(row=0, column=5, padx=4)
        ttk.Button(frm, text="⏰ Time", width=8, command=self._pick_time_for_selection).grid(row=0, column=6, padx=4)

        frm.columnconfigure(1, weight=1)

        # Distribution mode
        frm2 = ttk.Frame(self, padding=(10, 6, 10, 0))
        frm2.pack(fill=tk.X)
        ttk.Label(frm2, text="Distribution mode:").pack(side=tk.LEFT)
        ttk.Radiobutton(frm2, text="Repeat", variable=self.mode_var, value="titles").pack(side=tk.LEFT, padx=(8, 0))
        ttk.Radiobutton(frm2, text="No Repeat", variable=self.mode_var, value="channels").pack(side=tk.LEFT, padx=(8, 0))
        frm.columnconfigure(1, weight=1) 
    def _build_inputs(self):
        frm = ttk.Frame(self, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        left = ttk.Frame(frm)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        ttk.Label(left, text="Titles (one per line)").pack(anchor="w")
        self.txt_titles = tk.Text(left, height=12, wrap=tk.WORD)
        self.txt_titles.pack(fill=tk.BOTH, expand=True)

        right = ttk.Frame(frm)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        ttk.Label(right, text="Descriptions (1 line for all, or multiple lines)").pack(anchor="w")
        self.txt_descs = tk.Text(right, height=12, wrap=tk.WORD)
        self.txt_descs.pack(fill=tk.BOTH, expand=True)

        btns = ttk.Frame(self, padding=(10, 0, 10, 0))
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="Preview Distribution", command=self._preview).pack(side=tk.LEFT)
        ttk.Button(btns, text="Clear Inputs", command=self._clear_inputs).pack(side=tk.LEFT, padx=6)

    def _build_preview(self):
        frm = ttk.Frame(self, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        # THAY đổi: thêm 2 cột publish_date, publish_time
        cols = ("channel", "title", "description", "publish_date", "publish_time")
        self.tree = ttk.Treeview(frm, columns=cols, show="headings", height=12)
        for col in cols:
            self.tree.heading(col, text=col.capitalize())
            if col == "description":
                self.tree.column(col, width=420, anchor="w")
            elif col == "title":
                self.tree.column(col, width=300, anchor="w")
            elif col == "channel":
                self.tree.column(col, width=200, anchor="w")
            elif col == "publish_date":
                self.tree.column(col, width=120, anchor="w")
            elif col == "publish_time":
                self.tree.column(col, width=100, anchor="w")

        vsb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=vsb.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.LEFT, fill=tk.Y)

        # BIND double-click để mở hộp sửa hàng
        self.tree.bind("<Double-1>", self._on_tree_double_click)

        btns = ttk.Frame(self, padding=(10, 0, 10, 10))
        btns.pack(fill=tk.X)

        # nếu trước đó em đã đổi nút này thành Save Excel
        ttk.Button(btns, text="Save Excel", command=self._save_excel).pack(side=tk.LEFT)
        ttk.Button(btns, text="Copy TSV", command=self._copy_tsv).pack(side=tk.LEFT, padx=6)

        btns = ttk.Frame(self, padding=(10, 0, 10, 10))
        btns.pack(fill=tk.X)

        ttk.Button(btns, text="Save Excel", command=self._save_excel).pack(side=tk.LEFT)
        ttk.Button(btns, text="Copy TSV", command=self._copy_tsv).pack(side=tk.LEFT, padx=6)

        # THÊM 2 NÚT CHỌN DATE/TIME




    def _build_footer(self):
        bar = ttk.Frame(self, relief=tk.SUNKEN, padding=6)
        bar.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Label(bar, textvariable=self.status_var).pack(side=tk.LEFT)

    def _refresh_group_files(self):
        files = list_group_csvs(GROUPS_DIR)
        self.group_combo["values"] = files
        cur = self.group_file_var.get()
        if not files:
            self.group_file_var.set("")
            self.channel_count_lbl.config(text="0 channels")
            self._set_status(f"No CSV files in: {GROUPS_DIR}")
            return
        if cur not in files:
            self.group_file_var.set(files[0])
        self._load_channels()

    def _open_group_dir(self):
        if os.path.isdir(GROUPS_DIR):
            filedialog.askopenfilename(initialdir=GROUPS_DIR, title="(Read-only) group folder")
        else:
            messagebox.showwarning("Missing folder", f"'group' folder not found at:\n{GROUPS_DIR}")

    def _load_channels(self):
        name = self.group_file_var.get().strip()
        if not name:
            return
        csv_path = os.path.join(GROUPS_DIR, name)
        channels = read_channels_from_csv(csv_path)
        self._channels_cache = channels
        self.channel_count_lbl.config(text=f"{len(channels)} channels")
        self._set_status(f"Loaded {len(channels)} channels from {csv_path}")

    def _clear_inputs(self):
        self.txt_titles.delete("1.0", tk.END)
        self.txt_descs.delete("1.0", tk.END)
        self.tree.delete(*self.tree.get_children())
        self._last_assignments = None
        self._set_status("Cleared inputs & preview.")

    def _preview(self):
        group_file = self.group_file_var.get().strip()
        if not group_file:
            messagebox.showwarning("Missing CSV", "Please select a group CSV in ./group.")
            return

        titles = normalize_lines(self.txt_titles.get("1.0", tk.END))
        descs = normalize_lines(self.txt_descs.get("1.0", tk.END))
        channels = self._channels_cache
        mode = self.mode_var.get()

        try:
            assignments = assign_pairs(channels, titles, descs, mode=mode)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return

        self.tree.delete(*self.tree.get_children())
        extended = []
        for ch, t, d in assignments:
            pd, pt = "", "" 
            self.tree.insert("", tk.END, values=(ch, t, d, pd, pt))
            extended.append((ch, t, d, pd, pt))


        self._last_assignments = extended

        if mode == "titles":
            self._set_status(f"Previewed {len(assignments)} rows (title-driven; channels cycle if needed).")
        else:
            self._set_status(f"Previewed {len(assignments)} rows (channel-driven).")

    def _save_excel(self):
        if not self._last_assignments:
            messagebox.showwarning("Nothing to save", "Click Preview first.")
            return

        base = os.path.splitext(self.group_file_var.get().strip())[0] or "group"
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"{base}__assignments_{ts}.xlsx"

        os.makedirs(OUTPUT_DIR, exist_ok=True)
        out_path = os.path.join(OUTPUT_DIR, out_name)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Assignments"

            # THAY: thêm 2 cột publish_date, publish_time
            headers = ["group_file", "channel", "title", "description", "publish_date", "publish_time", "mode"]
            ws.append(headers)
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)

            mode_val = self.mode_var.get()
            group_file_shown = self.group_file_var.get()

            # Ghi từng dòng
            for ch, t, d, pd, pt in self._last_assignments:
                ws.append([group_file_shown, ch, t, d, pd, pt, mode_val])

            # Auto-fit tương đối
            desc_idx = headers.index("description") + 1
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                        min_col=col_idx, max_col=col_idx):
                    val = row[0].value
                    if val is None:
                        continue
                    s = str(val)
                    if col_idx == desc_idx:
                        s = s[:120]  # tránh quá rộng
                    max_len = max(max_len, len(s))

                # Gán width hợp lý
                header_name = headers[col_idx - 1]
                if header_name in ("publish_date", "publish_time"):
                    ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 18))
                else:
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            wb.save(out_path)

            self._set_status(f"Saved Excel: {out_path}")
            messagebox.showinfo("Saved", f"Exported Excel:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel:\n{e}")

    def _copy_tsv(self):
        if not self._last_assignments:
            messagebox.showwarning("Nothing to copy", "Click Preview first.")
            return
        # channel, title, description, publish_date, publish_time
        header = "channel\ttitle\tdescription\tpublish_date\tpublish_time\n"
        body = "\n".join(f"{ch}\t{t}\t{d}\t{pd}\t{pt}" for ch, t, d, pd, pt in self._last_assignments)
        tsv = header + body
        self.clipboard_clear()
        self.clipboard_append(tsv)
        self._set_status("Assignments copied to clipboard as TSV.")


    def _set_status(self, msg: str):
        self.status_var.set(msg)

    def _on_tree_double_click(self, event):
        # Xác định item được click (bỏ qua nếu click vào header/trống)
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return
        index = self.tree.index(item_id)  # vị trí trong tree (0-based)
        self._edit_row_dialog(item_id, index)

    def _edit_row_dialog(self, item_id, index):
        # Lấy giá trị hiện có
        vals = list(self.tree.item(item_id, "values"))
        if len(vals) < 5:
            # đảm bảo đủ 5 trường
            vals = list(vals) + [""] * (5 - len(vals))
        ch_cur, title_cur, desc_cur, pd_cur, pt_cur = vals

        win = tk.Toplevel(self)
        win.title("Edit row")
        win.transient(self)
        win.grab_set()

        frm = ttk.Frame(win, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        # Channel
        ttk.Label(frm, text="Channel:").grid(row=0, column=0, sticky="e", padx=6, pady=4)
        ent_ch = ttk.Entry(frm, width=60)
        ent_ch.grid(row=0, column=1, sticky="we")
        ent_ch.insert(0, ch_cur)

        # Title
        ttk.Label(frm, text="Title:").grid(row=1, column=0, sticky="e", padx=6, pady=4)
        ent_title = ttk.Entry(frm, width=60)
        ent_title.grid(row=1, column=1, sticky="we")
        ent_title.insert(0, title_cur)

        # Description (Text nhiều dòng)
        ttk.Label(frm, text="Description:").grid(row=2, column=0, sticky="ne", padx=6, pady=4)
        txt_desc = tk.Text(frm, width=60, height=6, wrap=tk.WORD)
        txt_desc.grid(row=2, column=1, sticky="we")
        txt_desc.insert("1.0", desc_cur)

        # Publish date (YYYY-MM-DD)
        ttk.Label(frm, text="Publish date (YYYY-MM-DD):").grid(row=3, column=0, sticky="e", padx=6, pady=4)
        ent_pd = ttk.Entry(frm, width=20)
        ent_pd.grid(row=3, column=1, sticky="w")
        ent_pd.insert(0, pd_cur)

        # Publish time (HH:MM, 24h)
        ttk.Label(frm, text="Publish time (HH:MM):").grid(row=4, column=0, sticky="e", padx=6, pady=4)
        ent_pt = ttk.Entry(frm, width=20)
        ent_pt.grid(row=4, column=1, sticky="w")
        ent_pt.insert(0, pt_cur)

        frm.columnconfigure(1, weight=1)

        def is_valid_date(s: str) -> bool:
            if not s.strip():
                return True  # cho phép bỏ trống
            try:
                datetime.datetime.strptime(s.strip(), "%Y-%m-%d")
                return True
            except ValueError:
                return False

        def is_valid_time(s: str) -> bool:
            if not s.strip():
                return True
            try:
                datetime.datetime.strptime(s.strip(), "%H:%M")
                return True
            except ValueError:
                return False

        def on_save():
            ch = ent_ch.get().strip()
            t = ent_title.get().strip()
            d = txt_desc.get("1.0", tk.END).strip()
            pd = ent_pd.get().strip()
            pt = ent_pt.get().strip()

            if not ch or not t:
                messagebox.showwarning("Missing", "Channel và Title không được để trống.")
                return
            if not is_valid_date(pd):
                messagebox.showerror("Invalid date", "Publish date phải dạng YYYY-MM-DD hoặc để trống.")
                return
            if not is_valid_time(pt):
                messagebox.showerror("Invalid time", "Publish time phải dạng HH:MM (24h) hoặc để trống.")
                return

            # Cập nhật Treeview
            new_vals = (ch, t, d, pd, pt)
            self.tree.item(item_id, values=new_vals)

            # Cập nhật _last_assignments tương ứng
            if 0 <= index < len(self._last_assignments):
                # _last_assignments: (ch, t, d, pd, pt)
                self._last_assignments[index] = new_vals

            self._set_status(f"Updated row {index+1}.")
            win.destroy()

        btns = ttk.Frame(win, padding=(0, 8))
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="Save", command=on_save).pack(side=tk.LEFT)
        ttk.Button(btns, text="Cancel", command=win.destroy).pack(side=tk.LEFT, padx=6)

        # Enter để lưu
        win.bind("<Return>", lambda e: on_save())
        # ESC để hủy
        win.bind("<Escape>", lambda e: win.destroy())

        # Đặt focus
        ent_title.focus_set()

    def _pick_date_for_selection(self):
        top = tk.Toplevel(self)
        top.title("Pick date for all rows")
        top.transient(self)
        top.grab_set()

        frm = ttk.Frame(top, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        today = datetime.date.today()

        if HAS_TKCAL:
            ttk.Label(frm, text="Chọn ngày:").grid(row=0, column=0, sticky="e", padx=4, pady=4)
            date_widget = DateEntry(frm, date_pattern="yyyy-mm-dd")
            date_widget.set_date(today)
            date_widget.grid(row=0, column=1, sticky="w")
        else:
            ttk.Label(frm, text="YYYY-MM-DD:").grid(row=0, column=0, sticky="e", padx=4, pady=4)
            date_widget = ttk.Entry(frm, width=14)
            date_widget.insert(0, today.strftime("%Y-%m-%d"))
            date_widget.grid(row=0, column=1, sticky="w")

        def ok():
            if HAS_TKCAL:
                d = date_widget.get_date()
                date_str = d.strftime("%Y-%m-%d")
            else:
                date_str = date_widget.get().strip()
                try:
                    datetime.datetime.strptime(date_str, "%Y-%m-%d")
                except ValueError:
                    messagebox.showerror("Invalid date", "Định dạng ngày phải là YYYY-MM-DD.")
                    return

            count = 0
            for iid in self.tree.get_children():
                idx = self.tree.index(iid)
                vals = list(self.tree.item(iid, "values"))
                vals += [""] * max(0, 5 - len(vals))
                ch, t, desc, pd, pt = vals
                new_vals = (ch, t, desc, date_str, pt)
                self.tree.item(iid, values=new_vals)
                if self._last_assignments and 0 <= idx < len(self._last_assignments):
                    self._last_assignments[idx] = new_vals
                count += 1

            self._set_status(f"Set publish_date='{date_str}' cho toàn bộ {count} dòng.")
            top.destroy()

        btns = ttk.Frame(top, padding=(0, 8))
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="OK", command=ok).pack(side=tk.LEFT)
        ttk.Button(btns, text="Cancel", command=top.destroy).pack(side=tk.LEFT, padx=6)



    def _pick_time_for_selection(self):
        top = tk.Toplevel(self)
        top.title("Pick time for all rows")
        top.transient(self)
        top.grab_set()

        frm = ttk.Frame(top, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        now = datetime.datetime.now()

        ttk.Label(frm, text="Giờ (00-23):").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        sp_h = tk.Spinbox(frm, from_=0, to=23, width=5, format="%02.0f")
        sp_h.delete(0, tk.END); sp_h.insert(0, f"{now.hour:02d}")
        sp_h.grid(row=0, column=1, sticky="w")

        ttk.Label(frm, text="Phút (00-59):").grid(row=1, column=0, sticky="e", padx=4, pady=4)
        sp_m = tk.Spinbox(frm, from_=0, to=59, width=5, format="%02.0f")
        sp_m.delete(0, tk.END); sp_m.insert(0, f"{now.minute:02d}")
        sp_m.grid(row=1, column=1, sticky="w")

        ttk.Label(frm, text="Step (phút):").grid(row=2, column=0, sticky="e", padx=4, pady=4)
        step_entry = ttk.Entry(frm, width=5)
        step_entry.insert(0, "0")  # 0 = giữ nguyên, >0 = tăng dần
        step_entry.grid(row=2, column=1, sticky="w")

        def ok():
            try:
                h = int(sp_h.get())
                m = int(sp_m.get())
                step = int(step_entry.get())
                assert 0 <= h < 24 and 0 <= m < 60 and step >= 0
            except:
                messagebox.showerror("Invalid input", "Vui lòng nhập giờ (0–23), phút (0–59), và step ≥ 0.")
                return

            base_time = datetime.datetime(2000, 1, 1, h, m)
            count = 0

            for i, iid in enumerate(self.tree.get_children()):
                idx = self.tree.index(iid)
                t_cur = base_time + datetime.timedelta(minutes=step * i)
                time_str = t_cur.strftime("%H:%M")

                vals = list(self.tree.item(iid, "values"))
                vals += [""] * max(0, 5 - len(vals))
                ch, t, desc, pd, pt = vals
                new_vals = (ch, t, desc, pd, time_str)
                self.tree.item(iid, values=new_vals)
                if self._last_assignments and 0 <= idx < len(self._last_assignments):
                    self._last_assignments[idx] = new_vals
                count += 1

            self._set_status(f"Set publish_time cho toàn bộ {count} dòng, bắt đầu từ {h:02d}:{m:02d}, step={step} phút.")
            top.destroy()

        btns = ttk.Frame(top, padding=(0, 8))
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="OK", command=ok).pack(side=tk.LEFT)
        ttk.Button(btns, text="Cancel", command=top.destroy).pack(side=tk.LEFT, padx=6)


