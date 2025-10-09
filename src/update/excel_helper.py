import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def save_assignments_to_excel(assignments, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignments"

    headers = ["channel", "directory", "title", "description", "publish_date", "publish_time"]
    ws.append(headers)
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)

    for row in assignments:
        ws.append(row)

    # Auto width
    desc_idx = headers.index("description") + 1
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is None:
                continue
            s = str(val)
            if col_idx == desc_idx:
                s = s[:120]
            max_len = max(max_len, len(s))

        header_name = headers[col_idx - 1]
        if header_name in ("publish_date", "publish_time"):
            ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 18))
        else:
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Ghi đè nếu tồn tại
    if os.path.exists(out_path):
        os.remove(out_path)

    wb.save(out_path)
    return out_path


def combine_excels(input_dir, output_file, move_folder, get_mp4_filename):
    import glob
    pattern = os.path.join(input_dir, "*.xlsx")
    files = glob.glob(pattern)
    if not files:
        return 0, []  # không có file

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Combined"

    header_written = False
    row_idx = 1
    processed_files = []

    for file in files:
        wb = load_workbook(file)
        ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column

        if not header_written:
            # copy header gốc
            for col in range(1, max_col + 1):
                ws_out.cell(row=1, column=col).value = ws.cell(row=1, column=col).value
            ws_out.cell(row=1, column=max_col + 1).value = "move_folder"
            header_written = True
            row_idx += 1

        for r in range(2, max_row + 1):
            row_values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]

            directory_val = row_values[1] if len(row_values) > 1 else ""
            filename = get_mp4_filename(directory_val)
            move_path = os.path.join(move_folder, filename) if filename else ""

            for c, val in enumerate(row_values, start=1):
                ws_out.cell(row=row_idx, column=c).value = val
            ws_out.cell(row=row_idx, column=max_col + 1).value = move_path

            row_idx += 1

        processed_files.append(file)

    # lưu file gộp
    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)
    if os.path.exists(output_file):
        os.remove(output_file)
    wb_out.save(output_file)

    # xóa dữ liệu trong file gốc
    for file in processed_files:
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None
        wb.save(file)

    return len(processed_files), processed_files